import uuid
import datetime
import getpass
import sys
import copy
from pprint import pprint as pp
import numpy as np
import pandas as pd
import re
import openpyxl as pyxl
import matplotlib.pyplot as plt
import atexit
from collections import defaultdict
import pyomrx

atexit.register(plt.show)

W_THICK = 2
W_DEFAULT = 1
W_THIN = 0.5
COLS = [pyxl.utils.cell.get_column_letter(i) for i in range(1, 99)]
ROW_HEIGHT_UNIT_PIXELS = 2
COLUMN_WIDTH_UNIT_PIXELS = 11
FONT_UNIT_PIXELS = 1
EMPTY_CIRCLE = b'\xe2\x97\x8b'
FULL_CIRCLE = b'\xe2\x97\x8f'
VERSION = pyomrx.__version__


def get_rectangle_from_range(range_cells, row_dims, col_dims):
    range_columns = [
        pyxl.utils.cell.get_column_letter(cell.column)
        for cell in range_cells[0]
    ]
    range_rows = [row[0].row for row in range_cells]
    form_heights = [row_dims[row]['ht'] for row in range_rows]
    form_widths = [col_dims[col]['width'] for col in range_columns]
    form_height = sum(form_heights)
    form_width = sum(form_widths)
    rows_above = []
    for row in range(1, 9999):
        if row in range_rows:
            break
        rows_above.append(row)
    cols_left = []
    height_above = sum([row_dims[row]['ht'] for row in rows_above])
    for col in COLS:
        if col in range_columns:
            break
        cols_left.append(col)
    width_left = sum([col_dims[col]['width'] for col in cols_left])
    return dict(
        top=-height_above,
        bottom=-(height_above + form_height),
        left=width_left,
        right=width_left + form_width)


def get_row_and_column_dimensions(form_sheet):
    raw_row_dims = [
        form_sheet.row_dimensions[row].__dict__ for row in range(1, 999)
    ]
    row_dims_df = pd.DataFrame(raw_row_dims).dropna()
    row_dims_df['ht'] *= ROW_HEIGHT_UNIT_PIXELS
    row_dims_df = row_dims_df.set_index('index')
    row_dims = row_dims_df.to_dict(orient='index')
    raw_col_dims = [form_sheet.column_dimensions[col].__dict__ for col in COLS]
    col_dims_df = pd.DataFrame(raw_col_dims).sort_values(by='index')
    col_dims_df['width'] = col_dims_df['width'].apply(lambda val: val if val
    else np.NaN).ffill() * COLUMN_WIDTH_UNIT_PIXELS
    col_dims_df = col_dims_df.set_index('index')
    col_dims = col_dims_df.to_dict(orient='index')
    return row_dims, col_dims


def get_relative_rectangle(inner_rectangle, outer_rectangle):
    outer_width = outer_rectangle['right'] - outer_rectangle['left']
    outer_height = outer_rectangle['top'] - outer_rectangle['bottom']
    relative_inner_rectangle = dict(top=(outer_rectangle['top'] - inner_rectangle['top']) / outer_height,
                                    bottom=(outer_rectangle['top'] - inner_rectangle['bottom']) / outer_height,
                                    left=(inner_rectangle['left'] - outer_rectangle['left']) / outer_width,
                                    right=(inner_rectangle['right'] - outer_rectangle['left']) / outer_width
                                    )
    return relative_inner_rectangle


class CircleParseError(Exception):
    pass


def parse_circles_from_range(metadata_range_cells, orient=None, assert_1d=False, allow_empty=False):
    metadata_values = []
    circle_radius = None
    for row in metadata_range_cells:
        row_values = []
        for cell in row:
            # print(f'{cell}: {cell.value}, {cell.font.sz}')
            encoded_cell_value = str(cell.value).encode()
            if encoded_cell_value == EMPTY_CIRCLE:
                row_values.append(False)
            elif encoded_cell_value == FULL_CIRCLE:
                row_values.append(True)
            elif encoded_cell_value == b'' and allow_empty:
                row_values.append(np.NaN)
            else:
                raise CircleParseError(f'character {cell.value} ({str(cell.value).encode()}) found, not allowed')
            if circle_radius is not None:
                if cell.font.sz * FONT_UNIT_PIXELS / 2 != circle_radius:
                    raise CircleParseError(f'varying font sizes, must be consistent')
            else:
                circle_radius = cell.font.sz * FONT_UNIT_PIXELS / 2
        metadata_values.append(row_values)
    metadata_arr = np.array(metadata_values)
    if assert_1d and not 1 in metadata_arr.shape:
        raise CircleParseError(f'2D circles area found: must be 1D')
    if metadata_arr.shape[1] > metadata_arr.shape[0]:
        metadata_orient = 'landscape'
    elif metadata_arr.shape[0] > metadata_arr.shape[1]:
        metadata_orient = 'portrait'
    else:
        metadata_orient = 'equal'
    if orient and metadata_orient != orient:
        raise CircleParseError('metadata field has shape {metadata_arr.shape}, '
                               f'so is {metadata_orient} but must be {orient}')
    metadata_arr = np.squeeze(metadata_arr)
    return dict(array=metadata_arr, orient=metadata_orient, radius=circle_radius)


def main():
    wb = pyxl.load_workbook(filename='temp/Absence register v28 (kc).xlsx', data_only=True)
    if 'template' not in wb:
        raise FileNotFoundError('couldnt find worksheet called "template"')
    form_sheet = wb['template']
    row_dims, col_dims = get_row_and_column_dimensions(form_sheet)
    range_names = [named_range.name for named_range in wb.get_named_ranges()]
    if 'page_1' not in range_names:
        raise ValueError('page_1 not found in named ranges')
    form_template_range = wb.get_named_range('page_1')
    print(form_template_range.__dict__)
    assert re.match(
        'template!',
        form_template_range.attr_text), 'form template range found in non template worksheet'
    range_cells = form_sheet[form_template_range.attr_text.split('!')[1]]
    form_rectangle = get_rectangle_from_range(
        range_cells=range_cells, row_dims=row_dims, col_dims=col_dims)
    form__template_fig, form_template_ax = plt.subplots(1, 1)
    form_template_ax.set_aspect('equal')
    form_top_left = dict(
        top=form_rectangle['top'], left=form_rectangle['left'])
    form_rectangle = localise_rectangle_to_form(form_rectangle,
                                                form_top_left)
    plot_rectangle(form_rectangle, form_template_ax, thickness=W_THICK)
    # TODO: asert that all sub forms are entirely inside page 1
    if 'sub_form_1' not in range_names:
        raise ValueError('page_1 not found in named ranges')
    sub_form_template_range = wb.get_named_range('sub_form_1')
    print(sub_form_template_range.__dict__)
    assert re.match(
        'template!',
        sub_form_template_range.attr_text), 'form template range found in non template worksheet'
    range_cells = form_sheet[sub_form_template_range.attr_text.split('!')[1]]
    sub_form_rectangle = get_rectangle_from_range(
        range_cells=range_cells, row_dims=row_dims, col_dims=col_dims)
    sub_form_rectangle = localise_rectangle_to_form(sub_form_rectangle,
                                                    form_top_left)
    plot_rectangle(sub_form_rectangle, form_template_ax, thickness=W_DEFAULT)

    metadata_ranges = [wb.get_named_range(range_name) for range_name in range_names if re.match('meta_', range_name)]
    metadata_circles_config = {}
    plt.figure()
    plt.xlim(0, 1)
    plt.ylim(1, 0)
    form_metadata_values = {}
    decides_sub_form_regex = 'decides sub form: (yes|no)'  # TODO: make more lenient
    # TODO: assert that metadata ranges are fully inside page 1
    for metadata_range in metadata_ranges:
        metadata_name = re.findall('meta_(.+)', metadata_range.name)[0]
        metadata_range_cells = form_sheet[metadata_range.attr_text.split('!')[1]]
        metadata_rectangle = get_rectangle_from_range(
            range_cells=metadata_range_cells, row_dims=row_dims, col_dims=col_dims)
        metadata_rectangle = localise_rectangle_to_form(metadata_rectangle,
                                                        form_top_left)
        metadata_relative_rectangle = get_relative_rectangle(metadata_rectangle, form_rectangle)
        plot_rectangle(metadata_relative_rectangle, thickness=W_THIN)
        plot_rectangle(metadata_rectangle, form_template_ax, thickness=W_THIN)
        metadata_circles_config[metadata_name] = metadata_relative_rectangle
        decides_sub_form = re.match(decides_sub_form_regex, str(metadata_range.comment))
        if decides_sub_form:
            if decides_sub_form == 'yes':
                decides_sub_form = True
            elif decides_sub_form == 'no':
                decides_sub_form = False
            else:
                raise ValueError(f'decide sub form is {decides_sub_form}, must be yes or no')
        else:
            print(f'WARNING: metadata field called {metadata_name} doesnt specify whether it decides the sub form, '
                  f'will assume that it does not (eg page number). '
                  f'To make the sub form depend on this metadata (eg exam type) add "decides sub form: yes"'
                  f'to the cell group\'s comment via the name manager')
            decides_sub_form = False
        metadata_circles_config[metadata_name]['decides_sub_form'] = decides_sub_form
        metadata_dict = parse_circles_from_range(metadata_range_cells,
                                                 orient='landscape',
                                                 assert_1d=True)
        metadata_arr = metadata_dict['array']
        circle_radius = metadata_dict['radius']
        metadata_value = 0
        for bit_index, cell_value in enumerate(reversed(metadata_arr.tolist())):
            metadata_value = metadata_value + bit_index ** 2 if cell_value else metadata_value
        if decides_sub_form:
            form_metadata_values[metadata_name] = metadata_value
        metadata_circles_config[metadata_name]['quantity'] = metadata_arr.size
        form_metadata_values[metadata_name] = metadata_value
        min_metadata_dimension = min([metadata_rectangle['top'] - metadata_rectangle['bottom'],
                                      metadata_rectangle['right'] - metadata_rectangle['left']])
        # print(f'rad:{circle_radius}, min metadata dimension:{min_metadata_dimension}')
        metadata_circles_config[metadata_name]['radius'] = circle_radius / min_metadata_dimension
        metadata_circle_offset = (metadata_rectangle['right'] - metadata_rectangle['left']) / metadata_arr.size
        left_circle_x = metadata_rectangle['left'] + metadata_circle_offset / 2
        left_circle_y = (metadata_rectangle['top'] + metadata_rectangle['bottom']) / 2
        circle_x_coords = [left_circle_x + metadata_circle_i * metadata_circle_offset for metadata_circle_i in
                           range(metadata_arr.size)]
        form_template_ax.plot(circle_x_coords, [left_circle_y] * metadata_arr.size,
                              'o')  # TODO: make these circles hollow and the correct size, might need shapely
    circles_ranges = [wb.get_named_range(range_name) for range_name in range_names if re.match('circles_', range_name)]
    if not circles_ranges:
        raise ValueError('no named ranges with the format cirlces_<name>')
    sub_form_template_config = dict(circles=dict(), metadata_requirements={})
    # TODO: make these regexes more lenient
    row_fill_regex = 'row fill: (many|one)'
    col_prefix_regex = 'column prefix: ([a-z|A-Z|0-9])'
    default_col_prefix = 'a'
    for circles_range in circles_ranges:
        circles_name = re.findall('circles_(.+)', circles_range.name)[0]
        circles_range_cells = form_sheet[circles_range.attr_text.split('!')[1]]
        circles_rectangle = get_rectangle_from_range(
            range_cells=circles_range_cells, row_dims=row_dims, col_dims=col_dims)
        circles_rectangle = localise_rectangle_to_form(circles_rectangle,
                                                       form_top_left)
        circles_rectangle_relative = get_relative_rectangle(sub_form_rectangle, form_rectangle)
        circles_config = copy.deepcopy(circles_rectangle_relative)
        plot_rectangle(circles_rectangle, form_template_ax, thickness=W_THIN)
        circles_dict = parse_circles_from_range(circles_range_cells)
        circles_arr = circles_dict['array']
        circle_offset_x = (circles_rectangle['right'] - circles_rectangle['left']) / circles_arr.shape[1]
        circle_offset_y = (circles_rectangle['top'] - circles_rectangle['bottom']) / circles_arr.shape[0]
        top_left_circle_x = circles_rectangle['left'] + circle_offset_x / 2
        top_left_circle_y = circles_rectangle['top'] - circle_offset_y / 2
        circle_x_coords = []
        circle_y_coords = []
        circles_per_row = []
        for row_i in range(circles_arr.shape[0]):
            circles_per_row.append(0)
            for col_i in range(circles_arr.shape[1]):
                if not np.isnan(circles_arr[row_i,col_i]):
                    circle_x_coords.append(top_left_circle_x + col_i * circle_offset_x)
                    circle_y_coords.append(top_left_circle_y - row_i * circle_offset_y)
                    circles_per_row[-1] += 1
        form_template_ax.plot(circle_x_coords, circle_y_coords, 'o')  # TODO: make these circles hollow and the correct size, might need shapely
        sub_form_template_config[circles_name] = circles_config
        sub_form_template_config['circles_per_row'] = circles_per_row
        circles_comment = str(circles_range.comment)
        row_fill = re.match(row_fill_regex, circles_comment)
        if row_fill:
            sub_form_template_config['allowed_row_filling'] = row_fill
        else:
            print(f'WARNING: row fill not specified for circles group called {circles_name}, '
                  f'will assume many circles per row can be filled in. To enforce one per row add "row fill: one"'
                  f'to the cell group\'s comment via the name manager')
        column_prefix = re.match(col_prefix_regex, circles_comment)
        if not column_prefix:
            column_prefix = default_col_prefix
            print(f'WARNING: column prefix not specified for circles group called {circles_name}, '
                  f'will prefix with {column_prefix}. To specifc a custom prefix add "column prefix: <prefix>"'
                  f'to the cell group\'s comment via the name manager')
            default_col_prefix = default_col_prefix + 'a'  # TODO: make this cycle the alphabet
        sub_form_template_config['column_prefix'] = column_prefix
        sub_form_template_config['radius'] = circles_dict['radius']
        sub_form_template_config['possible_rows'] = circles_arr.shape[0]
        sub_form_template_config['possible_columns'] = circles_arr.shape[1]
        # TODO: assert that all circles are completely inside the sub form
    sub_forms_config = dict(sub_form_templates=[sub_form_template_config],
                            locations=[circles_rectangle_relative])

    # TODO: print text from cells in plt

    # TODO: display cell borders as shown in excel

    # TODO: display cell shading as shown in excel

    # TODO: refactor OMR tool to use heirarchical form-subform-circles structure, with harder coding of locations
    # TODO: produce config for OMR tool automatically
    # TODO: save each permutation's excel file in a '.omrx' (actually a zip) archive along with the json config
    output_config = {}
    output_config['created_by'] = getpass.getuser()
    # output_config['description'] = input('form description: ')
    output_config['created_on'] = datetime.datetime.now()
    output_config['id'] = str(uuid.uuid4())
    template = dict(metadata_circles=metadata_circles_config, sub_forms=sub_forms_config)
    output_config['template'] = template
    output_config['pyomrx_version'] = VERSION
    pp(output_config)
    print('done')


def localise_rectangle_to_form(rectangle, parent_top_left, in_place=False):
    rectangle = rectangle if in_place else copy.deepcopy(rectangle)
    rectangle['top'] = rectangle['top'] - parent_top_left['top']
    rectangle['bottom'] = rectangle['bottom'] - parent_top_left['top']
    return rectangle


def plot_rectangle(rectangle, ax=None, thickness=W_DEFAULT):
    top = rectangle['top']
    left = rectangle['left']
    bottom = rectangle['bottom']
    right = rectangle['right']
    if ax:
        ax.plot([left, right, right, left, left], [top, top, bottom, bottom, top],
                c='black',
                linewidth=thickness)
    else:
        plt.plot([left, right, right, left, left], [top, top, bottom, bottom, top],
                 c='black',
                 linewidth=thickness)


if __name__ == '__main__':
    main()
