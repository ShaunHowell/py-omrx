# atexit.register(plt.show)
from pyomrx.core.exceptions import CircleParseError
from pyomrx.gui import FORM_GENERATION_TOPIC
from pubsub import pub
import shutil
from matplotlib.patches import Circle
from lxml import etree
from functools import reduce
from itertools import product
from pathlib import Path
import json
import os
import uuid
import datetime
import getpass
import sys
import copy
from pprint import pprint as pp
import numpy as np
import pandas as pd
import re
import openpyxl as pyxl
import matplotlib.pyplot as plt
import pyomrx
from pyomrx.core.meta import Abortable

LANDSCAPE = 'landscape'
PORTRAIT = 'portrait'
W_THICK = 5
W_DEFAULT = 1
W_THIN = 0.5
COLS = [pyxl.utils.cell.get_column_letter(i) for i in range(1, 99)]
ROW_HEIGHT_UNIT_PIXELS = 2
COLUMN_WIDTH_UNIT_PIXELS = 11
FONT_UNIT_PIXELS = 1.1
CIRCLE_FONT_MULTIPLIER = 0.4
EMPTY_CIRCLE = b'\xe2\x97\x8b'
FULL_CIRCLE = b'\xe2\x97\x8f'
COMMENT_PROP_START_REGEX = '(?:\A|\n|,\s*)'
COMMENT_PROP_END_REGEX = '([^,\n]+)'
BORDER_WIDTH_LOOKUP = dict(thin=0.5, medium=1, thick=2)
VERSION = pyomrx.__version__
X_TEXT_BUFFER = 1
ANY_NS_REGEX = '{.*}'
TEMP_FOLDER = Path('./pyomrxtemp')
VA_HA_MAP = dict(top='right', center='center', bottom='left')
HA_VA_MAP = dict(right='bottom', center='center', left='top')
CIRCLE_LINE_COLOUR = (0.8, 0.8, 0.8)


def get_rectangle_from_range(range_cells, row_dims, col_dims):
    range_columns = [
        pyxl.utils.cell.get_column_letter(cell.column)
        for cell in range_cells[0]
    ]
    range_rows = [row[0].row for row in range_cells]
    form_heights = [row_dims[row]['ht'] for row in range_rows]
    form_widths = [col_dims[col]['width'] for col in range_columns]
    form_height = sum(form_heights)
    form_width = sum(form_widths)
    rows_above = []
    for row in range(1, 9999):
        if row in range_rows:
            break
        rows_above.append(row)
    cols_left = []
    height_above = sum([row_dims[row]['ht'] for row in rows_above])
    for col in COLS:
        if col in range_columns:
            break
        cols_left.append(col)
    width_left = sum([col_dims[col]['width'] for col in cols_left])
    return dict(
        top=-height_above,
        bottom=-(height_above + form_height),
        left=width_left,
        right=width_left + form_width)


def assert_equal_rectangles(rect_a, rect_b, row_dims, col_dims, form_sheet):
    rect_a_cells = form_sheet[rect_a.split('!')[1]]
    rect_b_cells = form_sheet[rect_b.split('!')[1]]
    rect_a_rectangle = get_rectangle_from_range(rect_a_cells, row_dims,
                                                col_dims)
    rect_b_rectangle = get_rectangle_from_range(rect_b_cells, row_dims,
                                                col_dims)
    error_reason = ''
    area_a = get_rectangle_area(rect_a_rectangle)
    area_b = get_rectangle_area(rect_b_rectangle)
    if not np.isclose(area_a, area_b, atol=0.01):
        rect_a_rows = [row[0].row for row in rect_a_cells]
        rect_b_rows = [row[0].row for row in rect_b_cells]
        if len(rect_a_rows) != len(rect_b_rows):
            error_reason = f'{len(rect_a_rows)} in one form and {len(rect_b_rows)} in another'
        else:
            for row_a, row_b in zip(rect_a_rows, rect_b_rows):
                row_a_height = row_dims[row_a]['ht']
                row_b_height = row_dims[row_b]['ht']
                if row_a_height != row_b_height:
                    error_reason += f'{row_a} has height of {row_a_height} but ' \
                        f'{row_b} has height of {row_b_height}\n'
        raise ValueError(
            f'Range with cells {rect_b} does not have the same size as the form template.\n'
            f'a: {area_a}, b: {area_b}.\n'
            f'Please check all pages have identical row and column dimensions.\n'
            f'Row dimensions: {row_dims}.\n'
            f'Reason: {error_reason}')


def get_rectangle_area(rectangle):
    return (rectangle['top'] - rectangle['bottom']) * (
        rectangle['right'] - rectangle['left'])


def get_row_and_column_dimensions(form_sheet):
    raw_row_dims = [
        form_sheet.row_dimensions[row].__dict__ for row in range(1, 999)
    ]
    row_dims_df = pd.DataFrame(raw_row_dims).dropna()
    row_dims_df['ht'] *= ROW_HEIGHT_UNIT_PIXELS
    row_dims_df = row_dims_df.set_index('index')
    row_dims_df['top_y'] = -row_dims_df['ht'].cumsum().shift(1).fillna(0)
    row_dims = row_dims_df.to_dict(orient='index')
    raw_col_dims = [form_sheet.column_dimensions[col].__dict__ for col in COLS]
    col_dims_df = pd.DataFrame(raw_col_dims)
    col_dims_df['width'] = col_dims_df['width'].apply(
        lambda val: val if val else np.NaN).ffill() * COLUMN_WIDTH_UNIT_PIXELS
    col_dims_df['left_x'] = col_dims_df['width'].cumsum().shift(1).fillna(0)
    col_dims_df = col_dims_df.set_index('index')
    col_dims = col_dims_df.to_dict(orient='index')
    # pp(col_dims)
    return row_dims, col_dims


def get_relative_rectangle(inner_rectangle, outer_rectangle):
    outer_width = outer_rectangle['right'] - outer_rectangle['left']
    outer_height = outer_rectangle['top'] - outer_rectangle['bottom']
    relative_inner_rectangle = dict(
        top=(outer_rectangle['top'] - inner_rectangle['top']) / outer_height,
        bottom=(outer_rectangle['top'] - inner_rectangle['bottom']) /
        outer_height,
        left=(inner_rectangle['left'] - outer_rectangle['left']) / outer_width,
        right=(inner_rectangle['right'] - outer_rectangle['left']) /
        outer_width)
    return relative_inner_rectangle


def parse_circles_from_range(metadata_range_cells,
                             orient=None,
                             assert_1d=False,
                             allow_empty=False):
    metadata_values = []
    circle_radius = None
    for row in metadata_range_cells:
        row_values = []
        for cell in row:
            encoded_cell_value = str(cell.value).encode()
            if encoded_cell_value == EMPTY_CIRCLE:
                row_values.append(False)
            elif encoded_cell_value == FULL_CIRCLE:
                row_values.append(True)
            elif encoded_cell_value in [b'', b'None'] and allow_empty:
                row_values.append(np.NaN)
            else:
                raise CircleParseError(
                    f'character {cell.value} ({encoded_cell_value}) found in {cell}, not allowed'
                )
            if circle_radius is not None:
                if cell.font.sz * CIRCLE_FONT_MULTIPLIER * FONT_UNIT_PIXELS != circle_radius:
                    raise CircleParseError(
                        f'varying font sizes, must be consistent')
            else:
                circle_radius = cell.font.sz * CIRCLE_FONT_MULTIPLIER * FONT_UNIT_PIXELS
        metadata_values.append(row_values)
    circles_array = np.array(metadata_values)
    if assert_1d and not 1 in circles_array.shape:
        raise CircleParseError(f'2D circles area found: must be 1D')
    if circles_array.shape[1] > circles_array.shape[0]:
        orientation = LANDSCAPE
    elif circles_array.shape[0] > circles_array.shape[1]:
        orientation = PORTRAIT
    else:
        orientation = 'equal'
    if orient and orientation != orient:
        raise CircleParseError(
            f'circles group has shape {circles_array.shape}, '
            f'so is {orientation} but must be {orient}')
    # circles_array = np.squeeze(circles_array)
    return dict(array=circles_array, orient=orientation, radius=circle_radius)


def is_merged_cell(cell, merged_cells):
    return any(cell.coordinate in cells for cells in merged_cells)


def get_merged_cells(worksheet):
    merged_cells = []
    for merged_cell in worksheet.merged_cells.ranges:
        cells = []
        cols = [
            pyxl.utils.get_column_letter(i)
            for i in range(merged_cell.min_col, merged_cell.max_col + 1)
        ]
        rows = list(range(merged_cell.min_row, merged_cell.max_row + 1))
        for col, row in product(cols, rows):
            cells.append(f'{col}{row}')
        merged_cells.append(cells)
    return merged_cells


def assert_cells_contain_cells(outer_cells, inner_cells):
    assert cells_contain_cells(
        outer_cells,
        inner_cells), f'cells {inner_cells} not inside {outer_cells}'


def cells_contain_cells(outer_cells, inner_cells):
    assert '!' in outer_cells, f'outer cells {outer_cells} dont have worksheet prefix'
    assert '!' in inner_cells, f'inner cells {inner_cells} dont have worksheet prefix'
    outer_cells = copy.deepcopy(outer_cells).replace('$', '')
    inner_cells = copy.deepcopy(inner_cells).replace('$', '')
    inner_cells_sheet, inner_range_text = inner_cells.split('!')
    outer_cells_sheet, outer_range_text = outer_cells.split('!')
    if inner_cells_sheet == outer_cells_sheet:
        inner_first_col, inner_last_col = re.findall('([A-Z]+)[1-9][0-9]*',
                                                     inner_range_text)
        outer_first_col, outer_last_col = re.findall('([A-Z]+)[1-9][0-9]*',
                                                     outer_range_text)
        inner_first_col = pyxl.utils.column_index_from_string(inner_first_col)
        inner_last_col = pyxl.utils.column_index_from_string(inner_last_col)
        outer_first_col = pyxl.utils.column_index_from_string(outer_first_col)
        outer_last_col = pyxl.utils.column_index_from_string(outer_last_col)
        if inner_first_col >= outer_first_col and inner_last_col <= outer_last_col:
            inner_first_row, inner_last_row = re.findall(
                '[A-Z]+([1-9][0-9]*)', inner_range_text)
            outer_first_row, outer_last_row = re.findall(
                '[A-Z]+([1-9][0-9]*)', outer_range_text)
            if int(inner_first_row) >= int(outer_first_row) and int(
                    inner_last_row) <= int(outer_last_row):
                return True
    return False


def render_cell(ax, top, left, height, width, cell, theme_colours):
    if cell.border.left.style:
        border_thickness = BORDER_WIDTH_LOOKUP[cell.border.left.style]
        # print(f'border:{cell.border.left.color.__dict__}')
        border_colour = 'black'
        ax.plot([left, left], [top - height, top],
                c=border_colour,
                linewidth=border_thickness)
    if cell.border.right.style:
        border_thickness = BORDER_WIDTH_LOOKUP[cell.border.right.style]
        # print(f'border:{cell.border.right.color.__dict__}')
        border_colour = 'black'
        ax.plot([left + width, left + width], [top - height, top],
                c=border_colour,
                linewidth=border_thickness)
    if cell.border.top.style:
        border_thickness = BORDER_WIDTH_LOOKUP[cell.border.top.style]
        # print(f'border:{cell.border.top.color.__dict__}')
        border_colour = 'black'
        ax.plot([left, left + width], [top, top],
                c=border_colour,
                linewidth=border_thickness)
    if cell.border.bottom.style:
        border_thickness = BORDER_WIDTH_LOOKUP[cell.border.bottom.style]
        # print(f'border:{cell.border.bottom.color.__dict__}')
        border_colour = 'black'
        ax.plot([left, left + width], [top - height, top - height],
                c=border_colour,
                linewidth=border_thickness)
    if cell.fill.start_color.tint:
        rgb_val = 1 + cell.fill.start_color.tint
        ax.fill([left, left + width, left + width, left],
                [top, top, top - height, top - height],
                c=(rgb_val, rgb_val, rgb_val))
    if cell.value not in ['', None, np.NaN] and str(
            cell.value).encode() not in [FULL_CIRCLE, EMPTY_CIRCLE]:
        if cell.is_date:
            print(f'DATE CELL {cell} NOT RENDERED')
        else:
            # print(f'{cell} font colour: {cell.font.color.__dict__}')
            colour = '#000000'
            if cell.font.color is None:
                pass
            elif cell.font.color.type == 'theme':
                # super janky but have no idea why the theme colour list isn't sorted correctly
                if cell.font.color.theme == 1:
                    colour = theme_colours[0]
                elif cell.font.color.theme == 0:
                    colour = theme_colours[1]
                else:
                    print(f'couldnt parse font colour for cell {cell}')
            else:
                print(f'couldnt parse font colour for cell {cell}')
            font = dict(
                fontfamily=cell.font.name,
                fontsize=cell.font.sz * FONT_UNIT_PIXELS,
                fontstyle='italic' if cell.font.i else 'normal',
                fontweight='bold' if cell.font.b else 'normal',
            )
            # text_location
            ha = cell.alignment.horizontal or 'left'
            va = cell.alignment.vertical or 'center'
            if cell.alignment.textRotation == 0:
                x = left + FONT_UNIT_PIXELS * 4 if ha == 'left' else left + width / 2 if ha == 'center' \
                    else left + width - FONT_UNIT_PIXELS * 4
                y = top if va == 'top' else top - height / 2 if va == 'center' else top - height
                rotation = 0
            elif cell.alignment.textRotation == 180:
                # x = left + width if va == 'top' else left + width / 2 if va == 'center' else left
                # y = top - FONT_UNIT_PIXELS * 4 if ha == 'left' else top - height / 2 if ha == 'center' \
                #     else top - height + FONT_UNIT_PIXELS * 4
                # rotation = -90
                # old_ha = ha
                # ha = VA_HA_MAP[va]
                # va = HA_VA_MAP[old_ha]
                x = left + FONT_UNIT_PIXELS * 4 if ha == 'left' else left + width / 2 if ha == 'center' \
                    else left + width - FONT_UNIT_PIXELS * 4
                y = top if va == 'top' else top - height / 2 if va == 'center' else top - height
                rotation = -90
            else:
                raise ValueError(
                    f'cell rotation must either be 0 (horiontal) or 180 (?) (vertical). '
                    f'Got {cell.alignment.textRotation}')
            # print(dict(x=x, y=y,s=str(cell.value), fontdict=font,ha=ha, va=va))
            ax.text(
                x=x,
                y=y,
                s=str(cell.value),
                fontdict=font,
                ha=ha,
                va=va,
                color=colour,
                rotation=rotation)


def plot_circles(ax, centres, radius, fill):
    for centre in centres:
        circle = Circle((centre[0], centre[1]),
                        radius,
                        fill=fill,
                        edgecolor=CIRCLE_LINE_COLOUR,
                        facecolor='black')
        ax.add_patch(circle)


def get_circle_image_locations(circles_array, circles_rectangle):
    circle_offset_x = (circles_rectangle['right'] -
                       circles_rectangle['left']) / circles_array.shape[1]
    circle_offset_y = (circles_rectangle['top'] -
                       circles_rectangle['bottom']) / circles_array.shape[0]
    top_left_circle_x = circles_rectangle['left'] + circle_offset_x / 2
    top_left_circle_y = circles_rectangle['top'] - circle_offset_y / 2
    empty_circle_x_coords = []
    empty_circle_y_coords = []
    # filled_circle_x_coords = []
    # filled_circle_y_coords = []
    circles_per_row = []
    for row_i in range(circles_array.shape[0]):
        circles_per_row.append(0)
        for col_i in range(circles_array.shape[1]):
            if np.isnan(circles_array[row_i, col_i]):
                continue
            # elif circles_arr[row_i, col_i]:
            #     filled_circle_x_coords.append(top_left_circle_x +
            #                                         col_i * circle_offset_x)
            #     filled_circle_y_coords.append(top_left_circle_y -
            #                                   row_i * circle_offset_y)
            #     circles_per_row[-1] += 1
            elif not circles_array[row_i, col_i]:
                empty_circle_x_coords.append(top_left_circle_x +
                                             col_i * circle_offset_x)
                empty_circle_y_coords.append(top_left_circle_y -
                                             row_i * circle_offset_y)
                circles_per_row[-1] += 1
    return empty_circle_x_coords, empty_circle_y_coords, circles_per_row


def clean_temp_folder(temp_location, remake=False):
    if Path(temp_location).exists() and Path(temp_location).is_dir():
        shutil.rmtree(temp_location)
    if remake:
        os.makedirs(temp_location)


def get_theme_colours_from_wb(wb):
    theme_xml = etree.fromstring(wb.loaded_theme)
    # print(etree.tostring(theme_xml, pretty_print=True).decode())
    wb_theme_colours = []
    for top_el in filter(
            lambda el: re.match(ANY_NS_REGEX + 'themeElements', el.tag),
            theme_xml):
        # print(top_el.tag)
        for theme_el in filter(
                lambda el: re.match(ANY_NS_REGEX + 'clrScheme', el.tag),
                top_el):
            for colour_el in theme_el:
                colour_el = colour_el[0]
                if re.match(ANY_NS_REGEX + 'sysClr', colour_el.tag):
                    # print(f'sys: {colour_el.get("lastClr")}')
                    wb_theme_colours.append('#' + colour_el.get("lastClr"))
                elif re.match(ANY_NS_REGEX + 'srgbClr', colour_el.tag):
                    # print(f'srgb: {colour_el.get("val")}')
                    wb_theme_colours.append('#' + colour_el.get("val"))
                else:
                    print('theme colour not parsed:')
                    print(
                        etree.tostring(colour_el, pretty_print=True).decode())
    print(wb_theme_colours)
    return wb_theme_colours


def assert_circle_ranges_well_formed(circle_ranges):
    '''all circle ranges must be stacked perfectly in either a vetical or horizontal line,
     ie must all either share one x size or all share one y size'''

    assert False


def assert_range_is_single_rectangle(named_range):
    assert '+' not in named_range.attr_text,\
        f'named range {named_range.name} consists of more than one group of cells: {named_range.attr_text}'


class FormMaker(Abortable):
    def __init__(self,
                 excel_file_path,
                 output_path,
                 description=None,
                 abort_event=None,
                 id=None):
        Abortable.__init__(self, abort_event)
        print(f'generating form for excel file at {excel_file_path}')
        self.excel_file_path = Path(excel_file_path)
        self.output_path = Path(output_path)
        self.name = self.output_path.stem
        self.description = description or ''
        self.id = id
        self.wb = pyxl.load_workbook(
            filename=str(self.excel_file_path), data_only=True)
        if 'template' not in self.wb:
            raise FileNotFoundError('couldnt find worksheet called "template"')
        self.range_names = []
        form_component_keywords = ['page_', 'meta_', 'circles_', 'sub_form_']
        for named_range in self.wb.get_named_ranges():
            if any([
                    keyword in named_range.name
                    for keyword in form_component_keywords
            ]):
                assert_range_is_single_rectangle(named_range)
                self.range_names.append(named_range.name)
        self.row_dims, self.col_dims = get_row_and_column_dimensions(
            self.wb['template'])
        self.wb_theme_colours = get_theme_colours_from_wb(self.wb)
        page_name = 'page_1'
        if page_name not in self.range_names:
            raise ValueError(f'{page_name} not found in named ranges')
        self.form_template_range = self.wb.get_named_range(page_name)
        assert re.match(
            'template!', self.form_template_range.attr_text
        ), 'form template range found in non template worksheet'
        if 'sub_form_1' not in self.range_names:
            raise ValueError('sub_form_1 not found in named ranges')
        form_sheet = self.wb['template']
        self.sub_form_template_range = self.wb.get_named_range('sub_form_1')
        # self.sub_form_range_cells = form_sheet[sub_form_template_range.attr_text.split(
        #     '!')[1]]
        assert re.match(
            'template!', self.sub_form_template_range.attr_text
        ), 'form template range found in non template worksheet'

        self.metadata_ranges = [
            self.wb.get_named_range(range_name)
            for range_name in self.range_names
            if re.match('meta_', range_name)
        ]

        # for metadata_range in self.metadata_ranges:
        #     metadata_name = re.findall('meta_(.+)', metadata_range.name)[0]
        #     metadata_range_cells = form_sheet[metadata_range.attr_text.split(
        #         '!')[1]]

        self.circles_ranges = [
            self.wb.get_named_range(range_name)
            for range_name in self.range_names
            if re.match('circles_', range_name)
        ]
        if not self.circles_ranges:
            raise ValueError('no named ranges with the format cirlces_<name>')

    def update_progress(self, progress):
        if self.id:
            pub.sendMessage(
                f'{self.id}.{FORM_GENERATION_TOPIC}', progress=progress)

    def plot_form_page_image(self, page_number):
        page_name = f'page_{page_number}'
        if page_name not in self.range_names:
            raise ValueError(f'{page_name} not found in named ranges')
        page_range = self.wb.get_named_range(page_name)
        form_sheet = self.wb['template']
        form_template_range = self.translate_sub_range_to_new_parent(
            child=self.form_template_range.attr_text,
            old_parent=self.form_template_range.attr_text,
            new_parent=page_range.attr_text)
        form_template_range_cells = form_sheet[form_template_range.split('!')
                                               [1]]
        form_rectangle_abs = get_rectangle_from_range(
            range_cells=form_template_range_cells,
            row_dims=self.row_dims,
            col_dims=self.col_dims)
        assert_cells_contain_cells(page_range.attr_text, form_template_range)
        form_template_fig, form_template_ax = plt.subplots(1, 1, frameon=False)
        form_template_fig.set_size_inches(15.98, 11.93)  # A4
        form_template_ax.set_aspect('equal')
        form_top_left = dict(
            top=form_rectangle_abs['top'], left=form_rectangle_abs['left'])
        form_rectangle = localise_rectangle_to_form(form_rectangle_abs,
                                                    form_top_left)
        plot_rectangle(form_rectangle, form_template_ax, thickness=W_THICK)

        sub_form_template_range = self.translate_sub_range_to_new_parent(
            child=self.sub_form_template_range.attr_text,
            old_parent=self.form_template_range.attr_text,
            new_parent=page_range.attr_text)
        sub_form_template_cells = form_sheet[sub_form_template_range.split('!')
                                             [1]]
        assert_cells_contain_cells(form_template_range,
                                   sub_form_template_range)
        sub_form_template_rectangle = get_rectangle_from_range(
            range_cells=sub_form_template_cells,
            row_dims=self.row_dims,
            col_dims=self.col_dims)
        sub_form_template_rectangle = localise_rectangle_to_form(
            sub_form_template_rectangle, form_top_left)
        sub_form_template_rectangle_relative = get_relative_rectangle(
            sub_form_template_rectangle, form_rectangle)
        sub_form_locations = [sub_form_template_rectangle_relative]
        for sub_form_range_name in [
                name for name in self.range_names if 'sub_form_' in name
        ]:
            if sub_form_range_name == 'sub_form_1':
                continue
            sub_form_range = self.wb.get_named_range(sub_form_range_name)
            sub_form_range = self.translate_sub_range_to_new_parent(
                child=sub_form_range.attr_text,
                old_parent=self.form_template_range.attr_text,
                new_parent=page_range.attr_text)
            assert_cells_contain_cells(form_template_range, sub_form_range)
            sub_form_range_cells = form_sheet[sub_form_range.split('!')[1]]
            sub_form_rectangle = get_rectangle_from_range(
                range_cells=sub_form_range_cells,
                row_dims=self.row_dims,
                col_dims=self.col_dims)
            sub_form_rectangle = localise_rectangle_to_form(
                sub_form_rectangle, form_top_left)
            sub_form_rectangle_relative = get_relative_rectangle(
                sub_form_rectangle, form_rectangle)
            sub_form_locations.append(sub_form_rectangle_relative)
            for template_circles_range in self.circles_ranges:
                circles_range_str = self.translate_sub_range_to_new_parent(
                    child=template_circles_range.attr_text,
                    old_parent=self.sub_form_template_range.attr_text,
                    new_parent=sub_form_range)
                assert_cells_contain_cells(sub_form_range, circles_range_str),
                circles_name = re.findall('circles_(.+)',
                                          template_circles_range.name)[0]
                print(f'processing circles {circles_name}')
                circles_range_cells = form_sheet[circles_range_str.split('!')
                                                 [1]]
                circles_rectangle = get_rectangle_from_range(
                    range_cells=circles_range_cells,
                    row_dims=self.row_dims,
                    col_dims=self.col_dims)
                circles_rectangle = localise_rectangle_to_form(
                    circles_rectangle, form_top_left)
                circles_dict = parse_circles_from_range(
                    circles_range_cells, allow_empty=True)
                circles_arr = circles_dict['array']
                circle_x_coords, circle_y_coords, circles_per_row = get_circle_image_locations(
                    circles_arr, circles_rectangle)
                plot_circles(
                    form_template_ax,
                    zip(circle_x_coords, circle_y_coords),
                    circles_dict['radius'],
                    fill=False)
        form_metadata_circles_config = []
        self.raise_for_abort()
        form_metadata_values = {}
        for template_metadata_range in self.metadata_ranges:
            metadata_range_str = self.translate_sub_range_to_new_parent(
                child=template_metadata_range.attr_text,
                old_parent=self.form_template_range.attr_text,
                new_parent=page_range.attr_text)
            assert_cells_contain_cells(form_template_range, metadata_range_str)
            metadata_name = re.findall('meta_(.+)',
                                       template_metadata_range.name)[0]
            metadata_range_cells = form_sheet[metadata_range_str.split('!')[1]]
            metadata_rectangle = get_rectangle_from_range(
                range_cells=metadata_range_cells,
                row_dims=self.row_dims,
                col_dims=self.col_dims)
            metadata_rectangle = localise_rectangle_to_form(
                metadata_rectangle, form_top_left)
            metadata_relative_rectangle = get_relative_rectangle(
                metadata_rectangle, form_rectangle)
            # plot_rectangle(metadata_relative_rectangle, thickness=W_THIN)
            # plot_rectangle(metadata_rectangle, form_template_ax, thickness=W_THIN)
            metadata_circles_config = dict(
                rectangle=metadata_relative_rectangle, name=metadata_name)
            metadata_dict = parse_circles_from_range(
                metadata_range_cells, assert_1d=True)
            metadata_circles_config['orientation'] = metadata_dict['orient']
            metadata_arr = np.squeeze(metadata_dict['array'])
            circle_radius = metadata_dict['radius']
            metadata_value = 0
            for bit_index, cell_value in enumerate(
                    reversed(metadata_arr.tolist())):
                metadata_value = metadata_value + bit_index**2 if cell_value else metadata_value
            metadata_circles_config['quantity'] = metadata_arr.size
            form_metadata_values[metadata_name] = metadata_value
            max_metadata_dimension = max([
                metadata_rectangle['top'] - metadata_rectangle['bottom'],
                metadata_rectangle['right'] - metadata_rectangle['left']
            ])
            # print(f'rad:{circle_radius}, min metadata dimension:{min_metadata_dimension}')
            # print(f'metadata circle radius: {circle_radius}')
            # print(f'metadata circles rectangle max dim: {max_metadata_dimension}')
            # print(f'relative radius: {circle_radius / max_metadata_dimension}')
            metadata_circles_config[
                'radius'] = circle_radius / max_metadata_dimension
            if metadata_dict['orient'] == LANDSCAPE:
                metadata_circle_offset = (
                    metadata_rectangle['right'] -
                    metadata_rectangle['left']) / metadata_arr.size
                first_circle_x = metadata_rectangle[
                    'left'] + metadata_circle_offset / 2
                first_circle_y = (metadata_rectangle['top'] +
                                  metadata_rectangle['bottom']) / 2
            elif metadata_dict['orient'] == PORTRAIT:
                metadata_circle_offset = (
                    metadata_rectangle['top'] -
                    metadata_rectangle['bottom']) / metadata_arr.size
                first_circle_x = (metadata_rectangle['right'] +
                                  metadata_rectangle['left']) / 2
                first_circle_y = metadata_rectangle[
                    'top'] - metadata_circle_offset / 2

            empty_circle_varying_coords = []
            filled_circle_varying_coords = []
            for metadata_circle_i, circle_val in enumerate(metadata_arr):
                if metadata_dict['orient'] == LANDSCAPE:
                    unknown_circle_coord = first_circle_x + metadata_circle_i * metadata_circle_offset
                elif metadata_dict['orient'] == PORTRAIT:
                    unknown_circle_coord = first_circle_y - metadata_circle_i * metadata_circle_offset
                if np.isnan(circle_val):
                    continue
                elif circle_val:
                    filled_circle_varying_coords.append(unknown_circle_coord)
                else:
                    empty_circle_varying_coords.append(unknown_circle_coord)
            if metadata_dict['orient'] == LANDSCAPE:
                empty_circle_coords = zip(
                    empty_circle_varying_coords,
                    [first_circle_y] * len(empty_circle_varying_coords))
                filled_circle_coords = zip(
                    filled_circle_varying_coords,
                    [first_circle_y] * len(filled_circle_varying_coords))
            elif metadata_dict['orient'] == PORTRAIT:
                empty_circle_coords = zip(
                    [first_circle_x] * len(empty_circle_varying_coords),
                    empty_circle_varying_coords)
                filled_circle_coords = zip(
                    [first_circle_x] * len(filled_circle_varying_coords),
                    filled_circle_varying_coords)
            plot_circles(
                form_template_ax,
                empty_circle_coords,
                circle_radius,
                fill=False)
            plot_circles(
                form_template_ax,
                filled_circle_coords,
                circle_radius,
                fill=True)
            form_metadata_circles_config.append(metadata_circles_config)
        self.raise_for_abort()

        sub_form_template_config = dict(circles=[], metadata_requirements={})
        row_fill_regex = COMMENT_PROP_START_REGEX + 'row fill:\s*(many|one)'
        col_prefix_regex = COMMENT_PROP_START_REGEX + 'column prefix:\s*' + COMMENT_PROP_END_REGEX
        index_name_regex = COMMENT_PROP_START_REGEX + 'index name:\s*' + COMMENT_PROP_END_REGEX
        num_default_col_prefixes = 0
        shared_dimension = None
        reference_circles = None
        allowed_circles_fill_type = None
        for template_circles_range in self.circles_ranges:
            circles_range_str = self.translate_sub_range_to_new_parent(
                child=template_circles_range.attr_text,
                old_parent=self.form_template_range.attr_text,
                new_parent=page_range.attr_text)

            circles_name = re.findall('circles_(.+)',
                                      template_circles_range.name)[0]
            print(f'processing circles {circles_name}')
            circles_range_cells = form_sheet[circles_range_str.split('!')[1]]
            circles_rectangle = get_rectangle_from_range(
                range_cells=circles_range_cells,
                row_dims=self.row_dims,
                col_dims=self.col_dims)
            circles_rectangle = localise_rectangle_to_form(
                circles_rectangle, form_top_left)
            if reference_circles:
                if shared_dimension == 'x':
                    assert circles_rectangle['left'] == reference_circles[
                        'left']
                    assert circles_rectangle['right'] == reference_circles[
                        'right']
                elif shared_dimension == 'y':
                    assert circles_rectangle['top'] == reference_circles['top']
                    assert circles_rectangle['bottom'] == reference_circles[
                        'bottom']
                else:
                    if circles_rectangle['left'] == reference_circles['left'] and \
                            circles_rectangle['right'] == reference_circles['right']:
                        shared_dimension = 'x'
                    elif circles_rectangle['top'] == reference_circles['top'] and \
                            circles_rectangle['bottom'] == reference_circles['bottom']:
                        shared_dimension = 'y'
                    else:
                        raise ValueError(
                            'circle groups must all be either stacked vertically or horizontally, '
                            'and must be same width/height respectively')
            else:
                reference_circles = copy.deepcopy(circles_rectangle)
            # assert_circle_ranges_well_formed(self.circles_ranges)

            circles_rectangle_relative = get_relative_rectangle(
                circles_rectangle, sub_form_template_rectangle)
            circles_config = {
                'rectangle': copy.deepcopy(circles_rectangle_relative)
            }
            # plot_rectangle(circles_rectangle, form_template_ax, thickness=W_THIN)
            circles_dict = parse_circles_from_range(
                circles_range_cells, allow_empty=True)
            circles_arr = circles_dict['array']
            circle_x_coords, circle_y_coords, circles_per_row = get_circle_image_locations(
                circles_arr, circles_rectangle)
            plot_circles(
                form_template_ax,
                zip(circle_x_coords, circle_y_coords),
                circles_dict['radius'],
                fill=False)
            # plot_circles(
            #     form_template_ax,
            #     zip(filled_circle_x_coords, filled_circle_y_coords),
            #     circles_dict['radius'],
            #     fill=True)

            circles_comment = str(template_circles_range.comment)
            row_fill = re.findall(row_fill_regex, circles_comment)
            if not row_fill:
                raise ValueError(
                    f'row fill not specified for circles group called {circles_name}, '
                    f'cell groups comment should include "row fill: one," or "row_fill: many,"'
                )
            circles_config['allowed_row_filling'] = row_fill[0]
            if allowed_circles_fill_type:
                if allowed_circles_fill_type != circles_config[
                        'allowed_row_filling']:
                    raise ValueError(
                        'all circle groups must have same row fill type')
            else:
                allowed_circles_fill_type = circles_config[
                    'allowed_row_filling']
            column_prefix = re.findall(col_prefix_regex, circles_comment)
            if not column_prefix:
                num_default_col_prefixes += 1
                column_prefix = [
                    pyxl.utils.get_column_letter(num_default_col_prefixes)
                ]
                print(
                    f'WARNING: column prefix not specified for circles group called {circles_name}, '
                    f'will prefix with {column_prefix[0]}. To specifc a custom prefix add "column prefix: <prefix>" '
                    f'to the cell group\'s comment via the name manager')
            if column_prefix[0].lower() == 'none':
                column_prefix = None
            else:
                column_prefix = column_prefix[0]
            circles_config['column_prefix'] = column_prefix

            index_name = re.findall(index_name_regex, circles_comment)
            if index_name:
                circles_config['index_name'] = index_name[0]

            max_circles_dimension = max([
                circles_rectangle['top'] - circles_rectangle['bottom'],
                circles_rectangle['right'] - circles_rectangle['left']
            ])
            circles_config[
                'radius'] = circles_dict['radius'] / max_circles_dimension

            top_left_circle_cell = circles_range_cells[0][0]
            bottom_right_circle_cell = circles_range_cells[-1][-1]
            possible_rows = bottom_right_circle_cell.row - top_left_circle_cell.row + 1
            possible_columns = bottom_right_circle_cell.column - top_left_circle_cell.column + 1
            if circles_arr.shape[0] < possible_rows:
                circles_per_row.extend([0] * possible_rows -
                                       circles_arr.shape[0])
            circles_config['circles_per_row'] = circles_per_row
            circles_config['possible_columns'] = possible_columns
            circles_config['name'] = circles_name
            sub_form_template_config['circles'].append(circles_config)
        self.raise_for_abort()

        sub_forms_config = dict(
            sub_form_template=sub_form_template_config,
            locations=sub_form_locations)

        merged_cells = get_merged_cells(form_sheet)
        for row_index, row in enumerate(form_template_range_cells):
            row_top = self.row_dims[
                row[0].row]['top_y'] - form_rectangle_abs['top']
            row_height = self.row_dims[row[0].row]['ht']
            for column_index, cell in enumerate(row):
                if is_merged_cell(cell, merged_cells):
                    continue
                # print(f'{cell}: {cell.value}, {cell.font.sz}, {cell.row}:{cell.column}')
                column_left = self.col_dims[
                    cell.column_letter]['left_x'] - form_rectangle_abs['left']
                column_width = self.col_dims[cell.column_letter]['width']
                # form_template_ax.plot([column_left, column_left + column_width, column_left + column_width, column_left],
                #                       [row_top, row_top, row_top - row_height, row_top - row_height], c='black', alpha=0.05)
                render_cell(
                    form_template_ax,
                    row_top,
                    column_left,
                    row_height,
                    column_width,
                    cell,
                    theme_colours=self.wb_theme_colours)
                column_left += column_width
        print('finished parsing individual cells')
        print('parsing merged cells styles')
        self.raise_for_abort()
        for merged_cell in form_sheet.merged_cells.ranges:
            merged_cell_range_text = f'{form_sheet.title}!{merged_cell.coord}'
            if not cells_contain_cells(form_template_range,
                                       merged_cell_range_text):
                continue
            # print(f'plotting {merged_cell}')
            left_col = pyxl.utils.get_column_letter(merged_cell.min_col)
            top_left_cell = form_sheet[f'{left_col}{merged_cell.min_row}']
            merged_cell_left = self.col_dims[left_col][
                'left_x'] - form_rectangle_abs['left']
            merged_cell_width = reduce(
                lambda w, col_next: w + self.col_dims[
                    pyxl.utils.get_column_letter(col_next)]['width'],
                range(merged_cell.min_col, merged_cell.max_col + 1), 0)
            merged_cell_top = self.row_dims[
                merged_cell.min_row]['top_y'] - form_rectangle_abs['top']
            # print(merged_cell.min_row, merged_cell.max_row + 1)
            merged_cell_height = reduce(
                lambda h, row_next: h + self.row_dims[row_next]['ht'],
                range(merged_cell.min_row, merged_cell.max_row + 1), 0)
            # pp(dict(top=merged_cell_top, left=merged_cell_left, h=merged_cell_height, w=merged_cell_width,
            #         cell=top_left_cell,
            #         pt=(merged_cell.min_row == 1),
            #         pl=(merged_cell.min_col == 1)))
            render_cell(
                form_template_ax,
                merged_cell_top,
                merged_cell_left,
                merged_cell_height,
                merged_cell_width,
                top_left_cell,
                theme_colours=self.wb_theme_colours)
        self.raise_for_abort()
        template = dict(
            metadata_circles=form_metadata_circles_config,
            sub_forms=sub_forms_config)
        return form_template_fig, form_template_ax, template

    def translate_sub_range_to_new_parent(self, child, old_parent, new_parent):
        assert '!' in child, f'child doesnt have worksheet prefix: {child}'
        assert '!' in old_parent, f'old parent cells dont have worksheet prefix: {old_parent}'
        assert '!' in new_parent, f'new parent cells dont have worksheet prefix: {new_parent}'
        assert_cells_contain_cells(old_parent, child)
        form_sheet = self.wb['template']
        assert_equal_rectangles(old_parent, new_parent, self.row_dims,
                                self.col_dims, form_sheet)
        old_parent_range_sheet, old_parent_range_text = copy.deepcopy(
            old_parent).replace('$', '').split('!')
        new_parent_range_sheet, new_parent_range_text = copy.deepcopy(
            new_parent).replace('$', '').split('!')
        child_range_sheet, child_range_text = copy.deepcopy(child).replace(
            '$', '').split('!')

        if old_parent_range_sheet == new_parent_range_sheet:
            # translate columns
            old_parent_first_col, _ = re.findall('([A-Z]+)[1-9][0-9]*',
                                                 old_parent_range_text)
            new_parent_first_col, _ = re.findall('([A-Z]+)[1-9][0-9]*',
                                                 new_parent_range_text)
            child_first_col, child_last_col = re.findall(
                '([A-Z]+)[1-9][0-9]*', child_range_text)
            old_parent_first_col = pyxl.utils.column_index_from_string(
                old_parent_first_col)
            new_parent_first_col = pyxl.utils.column_index_from_string(
                new_parent_first_col)
            child_first_col = pyxl.utils.column_index_from_string(
                child_first_col)
            child_last_col = pyxl.utils.column_index_from_string(
                child_last_col)
            child_first_col_rel = child_first_col - old_parent_first_col
            new_child_first_col = new_parent_first_col + child_first_col_rel
            new_child_last_col = new_child_first_col + child_last_col - child_first_col
            new_child_first_col_letter = pyxl.utils.get_column_letter(
                new_child_first_col)
            new_child_last_col_letter = pyxl.utils.get_column_letter(
                new_child_last_col)

            # translate rows
            old_parent_first_row, _ = re.findall('[A-Z]+([1-9][0-9]*)',
                                                 old_parent_range_text)
            new_parent_first_row, _ = re.findall('[A-Z]+([1-9][0-9]*)',
                                                 new_parent_range_text)
            child_first_row, child_last_row = re.findall(
                '[A-Z]+([1-9][0-9]*)', child_range_text)
            child_first_row_rel = int(child_first_row) - int(
                old_parent_first_row)
            new_child_first_row = int(
                new_parent_first_row) + child_first_row_rel
            new_child_last_row = new_child_first_row + int(
                child_last_row) - int(child_first_row)

            return f'{child_range_sheet}!${new_child_first_col_letter}${new_child_first_row}:' \
                f'${new_child_last_col_letter}${new_child_last_row}'  # eg template!$D$3:$AH$27

        else:
            raise NotImplementedError(
                'template pages must all be on the same worksheet')

    def make_form(self):
        clean_temp_folder(TEMP_FOLDER, remake=True)
        filenames_to_copy = [self.excel_file_path]
        self.raise_for_abort()
        self.update_progress(0)
        description = self.description
        name = self.name
        form_template_fig, form_template_ax, template = self.plot_form_page_image(
            page_number=1)
        # for page_range_name in filter(lambda range_name: re.match('(page_[2-9][0-9]*)|(page_1[0-9]+)',
        #                                                           range_name), self.wb.get_named_ranges()):
        #     # plot the exact same circles, but filled correctly, and the correct text for the new range
        #     pass
        output_config = {}
        output_config['author'] = getpass.getuser()
        output_config['description'] = description
        output_config['name'] = name
        output_config['created_on'] = str(datetime.datetime.now())
        output_config['id'] = str(uuid.uuid4())
        output_config['template'] = template
        output_config['pyomrx_version'] = VERSION
        # pp(output_config)
        # output_folder = Path('pyomrx/tests/res/form_config')
        output_folder = self.output_path.parent
        os.makedirs(str(output_folder), exist_ok=True)
        config_temp_path = str(TEMP_FOLDER / 'omr_config.json')
        # pp(output_config)
        json.dump(output_config, open(config_temp_path, 'w'))
        strip_ax_padding(form_template_ax)
        form_template_fig.subplots_adjust(
            top=1, bottom=0, right=1, left=0, hspace=0, wspace=0)
        print('saving figure')
        page_1_form_temp_path = str(output_folder / f'{self.name}_page_1.png')
        filenames_to_copy.append(page_1_form_temp_path)
        form_template_fig.savefig(
            page_1_form_temp_path, bbox_inches='tight', pad_inches=0)
        # xlim = form_template_ax.get_xlim()
        # ylim = form_template_ax.get_ylim()
        # form_template_ax.set_xlim(xlim[0] - 10, xlim[1] + 10)
        # form_template_ax.set_ylim(ylim[0] - 10, ylim[1] + 10)
        # form_template_fig.savefig(
        #     str(output_folder / f'{self.excel_file_path.stem}.png'),
        #     bbox_inches='tight',
        #     pad_inches=0)

        # other pages
        page_numbers = [
            int(re.findall('page_([0-9]+)', name)[0])
            for name in self.range_names if re.match('page_[0-9]+', name)
        ]
        num_pages = len(page_numbers)
        self.update_progress(80 / num_pages)
        for i, page_number in enumerate(page_numbers):
            self.raise_for_abort()
            if page_number == 1:
                continue
            page_fig, page_ax, _ = self.plot_form_page_image(page_number)
            strip_ax_padding(page_ax)
            page_fig.subplots_adjust(
                top=1, bottom=0, right=1, left=0, hspace=0, wspace=0)
            print('saving figure')
            page_fig_temp_path = str(
                output_folder / f'{self.name}_page_{page_number}.png')
            filenames_to_copy.append(page_fig_temp_path)
            # plt.show()
            page_fig.savefig(
                page_fig_temp_path, bbox_inches='tight', pad_inches=0)
            self.update_progress((i + 1) * 80 / num_pages)

        print('making archive')
        for file_path in filenames_to_copy:
            file_path = Path(file_path)
            shutil.copy(
                file_path,
                str(TEMP_FOLDER / f'{file_path.stem}.{file_path.suffix}'))
        shutil.make_archive(str(output_folder / name), 'zip', str(TEMP_FOLDER))
        self.raise_for_abort()
        self.update_progress(90)
        if (output_folder / f'{name}.omr').exists():
            print(
                f'WARNING: {output_folder} / {name}.omr already exists, will overwrite'
            )
            os.remove(str(output_folder / f'{name}.omr'))
        os.rename(
            str(output_folder / f'{name}.zip'),
            str(output_folder / f'{name}.omr'))
        clean_temp_folder(TEMP_FOLDER, remake=False)
        self.update_progress(100)
        print('done')
        return output_config


def strip_ax_padding(ax, margin=0.02):
    ax.set_axis_off()
    ax.margins(margin, margin)
    ax.xaxis.set_major_locator(plt.NullLocator())
    ax.yaxis.set_major_locator(plt.NullLocator())
    xlim = ax.get_xlim()
    ylim = ax.get_ylim()
    ax.set_xlim(xlim[0] - W_THICK / 2, xlim[1] + W_THICK / 2)
    ax.set_ylim(ylim[0] - W_THICK / 2, ylim[1] + W_THICK / 2)


def localise_rectangle_to_form(rectangle, parent_top_left, in_place=False):
    rectangle = rectangle if in_place else copy.deepcopy(rectangle)
    rectangle['top'] = rectangle['top'] - parent_top_left['top']
    rectangle['bottom'] = rectangle['bottom'] - parent_top_left['top']
    return rectangle


def plot_rectangle(rectangle, ax=None, thickness=W_DEFAULT):
    top = rectangle['top']
    left = rectangle['left']
    bottom = rectangle['bottom']
    right = rectangle['right']
    if ax:
        ax.plot([left, right, right, left, left],
                [top, top, bottom, bottom, top],
                c='black',
                linewidth=thickness)
    else:
        plt.plot([left, right, right, left, left],
                 [top, top, bottom, bottom, top],
                 c='black',
                 linewidth=thickness)
